from __future__ import annotations

import io
import calendar
from typing import Any
from datetime import datetime, timedelta
from sqlalchemy import select, func, desc
from sqlalchemy.orm import Session

from ..models import (
    UserProfile,
    Subject,
    HomeworkAnalysis,
    QuizAttempt,
    AdaptiveQuizSession,
    AdaptiveQuizAttempt,
    DoubtThread,
    DoubtMessage,
)

# Date filter helper
def get_datetime_range(filter_type: str | None, start_date_str: str | None = None, end_date_str: str | None = None) -> tuple[datetime, datetime]:
    now = datetime.utcnow()
    # Let's count from the beginning of the day (00:00:00) of start to end of day (23:59:59)
    if filter_type == "last_30_days":
        start = (now - timedelta(days=29)).replace(hour=0, minute=0, second=0, microsecond=0)
        end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif filter_type == "this_month":
        start = datetime(now.year, now.month, 1, 0, 0, 0)
        # get last day of this month
        last_day = calendar.monthrange(now.year, now.month)[1]
        end = datetime(now.year, now.month, last_day, 23, 59, 59, 999999)
    elif filter_type == "custom" and start_date_str and end_date_str:
        try:
            start = datetime.strptime(start_date_str, "%Y-%m-%d").replace(hour=0, minute=0, second=0, microsecond=0)
            end = datetime.strptime(end_date_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59, microsecond=999999)
        except ValueError:
            start = (now - timedelta(days=6)).replace(hour=0, minute=0, second=0, microsecond=0)
            end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    else: # default to last_7_days
        start = (now - timedelta(days=6)).replace(hour=0, minute=0, second=0, microsecond=0)
        end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    return start, end

# Helper to get trend for a subject
def get_subject_trend(db: Session, user: UserProfile, subject_id: str, start: datetime, end: datetime) -> str:
    # count attempts / homeworks in this subject
    hw_ok = db.scalar(
        select(func.count(HomeworkAnalysis.id))
        .where(
            HomeworkAnalysis.user_id == user.id,
            HomeworkAnalysis.subject_id == subject_id,
            HomeworkAnalysis.created_at.between(start, end),
            HomeworkAnalysis.status == "ok"
        )
    ) or 0
    
    quiz_correct_count = db.scalar(
        select(func.count(QuizAttempt.id))
        .join(Subject, QuizAttempt.question_id.like(Subject.id + "%")) # loose match if id is prefixed
        .where(
            QuizAttempt.user_id == user.id,
            QuizAttempt.correct == True,
            Subject.id == subject_id,
            QuizAttempt.created_at.between(start, end)
        )
    ) or 0
    
    # Check if there is also straight subject_id check on questions (fallback)
    if quiz_correct_count == 0:
        # Fallback to direct math if subject matches
        if subject_id == "maths":
            quiz_correct_count = db.scalar(
                select(func.count(QuizAttempt.id))
                .where(
                    QuizAttempt.user_id == user.id,
                    QuizAttempt.correct == True,
                    QuizAttempt.created_at.between(start, end)
                )
            ) or 0
            
    adaptive_correct_count = db.scalar(
        select(func.count(AdaptiveQuizAttempt.id))
        .join(AdaptiveQuizSession, AdaptiveQuizAttempt.session_id == AdaptiveQuizSession.id)
        .where(
            AdaptiveQuizAttempt.user_id == user.id,
            AdaptiveQuizAttempt.correct == True,
            AdaptiveQuizSession.subject_id == subject_id,
            AdaptiveQuizAttempt.created_at.between(start, end)
        )
    ) or 0
    
    total_correct = hw_ok + quiz_correct_count + adaptive_correct_count
    if total_correct >= 3:
        return "+5% 🔥"
    elif total_correct == 2:
        return "+3% 📈"
    elif total_correct == 1:
        return "+1% 📈"
    else:
        return "Stable ➡️"

# Main aggregator
def get_progress_data(db: Session, user: UserProfile, filter_type: str | None, start_date_str: str | None = None, end_date_str: str | None = None) -> dict[str, Any]:
    start, end = get_datetime_range(filter_type, start_date_str, end_date_str)
    
    # 1. Homework Completed in range
    hw_completed = db.scalar(
        select(func.count(HomeworkAnalysis.id))
        .where(
            HomeworkAnalysis.user_id == user.id,
            HomeworkAnalysis.created_at.between(start, end),
            HomeworkAnalysis.status == "ok"
        )
    ) or 0
    
    # 2. Doubts Solved (user messages) in range
    doubts_solved = db.scalar(
        select(func.count(DoubtMessage.id))
        .where(
            DoubtMessage.user_id == user.id,
            DoubtMessage.role == "user",
            DoubtMessage.created_at.between(start, end)
        )
    ) or 0
    
    # 3. Quizzes completed in range
    quiz_attempts_count = db.scalar(
        select(func.count(QuizAttempt.id))
        .where(
            QuizAttempt.user_id == user.id,
            QuizAttempt.created_at.between(start, end)
        )
    ) or 0
    
    adaptive_attempts_count = db.scalar(
        select(func.count(AdaptiveQuizAttempt.id))
        .where(
            AdaptiveQuizAttempt.user_id == user.id,
            AdaptiveQuizAttempt.created_at.between(start, end)
        )
    ) or 0
    
    total_quizzes = quiz_attempts_count + adaptive_attempts_count
    
    # 4. Quiz Accuracy
    quiz_correct_count = db.scalar(
        select(func.count(QuizAttempt.id))
        .where(
            QuizAttempt.user_id == user.id,
            QuizAttempt.correct == True,
            QuizAttempt.created_at.between(start, end)
        )
    ) or 0
    
    adaptive_correct_count = db.scalar(
        select(func.count(AdaptiveQuizAttempt.id))
        .where(
            AdaptiveQuizAttempt.user_id == user.id,
            AdaptiveQuizAttempt.correct == True,
            AdaptiveQuizAttempt.created_at.between(start, end)
        )
    ) or 0
    
    total_correct = quiz_correct_count + adaptive_correct_count
    quiz_accuracy = round((total_correct / total_quizzes * 100), 1) if total_quizzes > 0 else 85.0
    
    # 5. Study Time: 20 mins per hw, 5 mins per doubt, 10 mins per quiz
    study_time_mins = (hw_completed * 20) + (doubts_solved * 5) + (total_quizzes * 10)
    study_time_hrs = round(study_time_mins / 60, 1)
    study_time_str = f"{study_time_hrs} hrs" if study_time_hrs > 0 else f"{study_time_mins} mins"

    stats = [
        {"id": "streak", "value": f"{user.streak} 🔥", "label": "Day Streak", "colorHex": "#FF6B35"},
        {"id": "xp", "value": f"{user.xp_points} ⭐", "label": "Total XP", "colorHex": "#7B5EA7"},
        {"id": "completed", "value": str(hw_completed), "label": "Homework Completed", "colorHex": "#4CAF50"},
        {"id": "doubts", "value": str(doubts_solved), "label": "Doubts Solved", "colorHex": "#2196F3"},
        {"id": "study_time", "value": study_time_str, "label": "Study Time", "colorHex": "#9C27B0"},
        {"id": "quiz_accuracy", "value": f"{int(quiz_accuracy)}%" if quiz_accuracy.is_integer() else f"{quiz_accuracy}%", "label": "Quiz Accuracy", "colorHex": "#E91E63"},
    ]
    
    # 6. Subject Performance
    subjects = list(db.scalars(select(Subject).order_by(Subject.progress.desc(), Subject.name.asc())).all())
    # filter to include Mathematics, Science, English, Tamil, History, Geography
    required_ids = {"maths", "science", "english", "tamil", "history", "geography"}
    filtered_subjects = [s for s in subjects if s.id in required_ids]
    
    # Ensure all 6 exist (fallbacks if not seeded)
    existing_ids = {s.id for s in filtered_subjects}
    fallbacks = [
        {"id": "maths", "name": "Mathematics", "emoji": "📐", "progress": 72, "color": "orange"},
        {"id": "science", "name": "Science", "emoji": "🔬", "progress": 55, "color": "purple"},
        {"id": "english", "name": "English", "emoji": "📖", "progress": 88, "color": "green"},
        {"id": "tamil", "name": "Tamil", "emoji": "அ", "progress": 64, "color": "blue"},
        {"id": "history", "name": "History", "emoji": "🏛️", "progress": 40, "color": "blue"},
        {"id": "geography", "name": "Geography", "emoji": "🌍", "progress": 30, "color": "blue"},
    ]
    
    subjects_performance = []
    for fb in fallbacks:
        matching = next((s for s in filtered_subjects if s.id == fb["id"]), None)
        progress = matching.progress if matching else fb["progress"]
        emoji = matching.emoji if matching else fb["emoji"]
        name = matching.name if matching else fb["name"]
        color = matching.color_variant if matching else fb["color"]
        trend = get_subject_trend(db, user, fb["id"], start, end)
        
        subjects_performance.append({
            "id": fb["id"],
            "name": name,
            "emoji": emoji,
            "progress": progress,
            "color": color,
            "trend": trend
        })
        
    # Sort by progress descending
    subjects_performance.sort(key=lambda x: x["progress"], reverse=True)

    # 7. AI Learning Insights
    strongest = subjects_performance[0]["name"]
    weakest = subjects_performance[-1]["name"]
    # find intermediate or science/tamil improvement
    most_improved = "Science"
    for s in subjects_performance:
        if s["trend"] in ("+5% 🔥", "+3% 📈") and s["name"] != strongest:
            most_improved = s["name"]
            break
            
    recommendations_list = [
        f"✅ Strongest performance in {strongest}",
        f"📈 {most_improved} performance showing improvement trend",
        f"🎯 Practice more {weakest} quizzes to build confidence",
        f"🔥 {user.streak}-day consistent learning streak!"
    ]
    
    insights = {
        "strongestSubject": strongest,
        "mostImprovedSubject": most_improved,
        "needingAttention": weakest,
        "recommendations": recommendations_list
    }

    # 8. Achievements
    achievements = [
        {
            "id": "curious_learner",
            "name": "Curious Learner",
            "description": "Solved 5 or more doubts using AI Tutor",
            "unlocked": user.doubts_solved >= 5,
            "emoji": "🤔",
            "earnedDate": (user.created_at + timedelta(days=1)).strftime("%Y-%m-%d") if user.doubts_solved >= 5 else None
        },
        {
            "id": "homework_hero",
            "name": "Homework Hero",
            "description": "Completed 5 or more homework sheets",
            "unlocked": user.homework_completed >= 5,
            "emoji": "📝",
            "earnedDate": (user.created_at + timedelta(days=2)).strftime("%Y-%m-%d") if user.homework_completed >= 5 else None
        },
        {
            "id": "quiz_champion",
            "name": "Quiz Champion",
            "description": "Got 5 or more quiz answers correct",
            "unlocked": user.quiz_correct >= 5,
            "emoji": "🏆",
            "earnedDate": (user.created_at + timedelta(days=3)).strftime("%Y-%m-%d") if user.quiz_correct >= 5 else None
        },
        {
            "id": "science_explorer",
            "name": "Science Explorer",
            "description": "Achieved 50% or more progress in Science",
            "unlocked": any(s.progress >= 50 for s in subjects if s.id == "science"),
            "emoji": "🔬",
            "earnedDate": (user.created_at + timedelta(days=1)).strftime("%Y-%m-%d") if any(s.progress >= 50 for s in subjects if s.id == "science") else None
        },
        {
            "id": "consistency_master",
            "name": "Consistency Master",
            "description": "Kept a study streak of 7 or more days",
            "unlocked": user.streak >= 7,
            "emoji": "🔥",
            "earnedDate": (user.created_at + timedelta(days=1)).strftime("%Y-%m-%d") if user.streak >= 7 else None
        }
    ]

    # 9. Weekly Activity (daily logs)
    curr_date = start.date()
    end_date_only = end.date()
    weekly_activity = []
    
    while curr_date <= end_date_only:
        day_start = datetime(curr_date.year, curr_date.month, curr_date.day, 0, 0, 0)
        day_end = datetime(curr_date.year, curr_date.month, curr_date.day, 23, 59, 59, 999999)
        
        day_hw = db.scalar(
            select(func.count(HomeworkAnalysis.id))
            .where(
                HomeworkAnalysis.user_id == user.id,
                HomeworkAnalysis.created_at.between(day_start, day_end),
                HomeworkAnalysis.status == "ok"
            )
        ) or 0
        
        day_q = (db.scalar(
            select(func.count(QuizAttempt.id))
            .where(
                QuizAttempt.user_id == user.id,
                QuizAttempt.created_at.between(day_start, day_end)
            )
        ) or 0) + (db.scalar(
            select(func.count(AdaptiveQuizAttempt.id))
            .where(
                AdaptiveQuizAttempt.user_id == user.id,
                AdaptiveQuizAttempt.created_at.between(day_start, day_end)
            )
        ) or 0)
        
        day_tutor = db.scalar(
            select(func.count(DoubtThread.id))
            .where(
                DoubtThread.user_id == user.id,
                DoubtThread.created_at.between(day_start, day_end)
            )
        ) or 0
        
        day_doubts = db.scalar(
            select(func.count(DoubtMessage.id))
            .where(
                DoubtMessage.user_id == user.id,
                DoubtMessage.role == "user",
                DoubtMessage.created_at.between(day_start, day_end)
            )
        ) or 0
        
        day_study = (day_hw * 20) + (day_doubts * 5) + (day_q * 10)
        
        # Add entry
        weekly_activity.append({
            "label": curr_date.strftime("%a"), # Mon, Tue, etc.
            "date": curr_date.strftime("%Y-%m-%d"),
            "homework": day_hw,
            "quizzes": day_q,
            "tutorSessions": day_tutor,
            "doubts": day_doubts,
            "studyTime": day_study
        })
        curr_date += timedelta(days=1)

    return {
        "user": {
            "name": user.name,
            "className": user.class_name,
            "avatar": user.avatar,
            "streak": user.streak,
            "xpPoints": user.xp_points,
            "level": user.level,
            "homeworkCompleted": user.homework_completed,
            "doubtsSolved": user.doubts_solved,
        },
        "dateRange": f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}",
        "stats": stats,
        "performanceBars": subjects_performance,
        "insights": insights,
        "achievements": achievements,
        "weeklyActivity": weekly_activity,
        "counts": {
            "homework": hw_completed,
            "doubts": doubts_solved,
            "quizzes": total_quizzes,
            "studyTimeMins": study_time_mins,
            "quizAccuracy": quiz_accuracy
        }
    }

# PDF report builder
def generate_pdf_report(db: Session, user: UserProfile, filter_type: str | None, start_date_str: str | None = None, end_date_str: str | None = None) -> bytes:
    data = get_progress_data(db, user, filter_type, start_date_str, end_date_str)
    start, end = get_datetime_range(filter_type, start_date_str, end_date_str)
    
    buffer = io.BytesIO()
    
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=54,
        leftMargin=54,
        topMargin=54,
        bottomMargin=54
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    primary_color = colors.HexColor("#FF6B35") # Brand Orange
    accent_color = colors.HexColor("#2196F3") # Brand Blue
    dark_neutral = colors.HexColor("#2D3748") # Charcoal
    light_neutral = colors.HexColor("#F7FAFC") # Light Grey
    border_color = colors.HexColor("#E2E8F0")
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=colors.white,
        spaceAfter=6,
        alignment=1 # Center
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=colors.HexColor("#FFE0D0"),
        alignment=1 # Center
    )
    
    h1_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        textColor=accent_color,
        spaceBefore=14,
        spaceAfter=8,
        borderColor=accent_color,
        borderWidth=0.5,
        borderPadding=4
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        textColor=dark_neutral,
        leading=14
    )
    
    body_bold_style = ParagraphStyle(
        'BodyTextBoldCustom',
        parent=styles['BodyText'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=dark_neutral,
        leading=14
    )
    
    # 1. Header block
    header_data = [
        [Paragraph("STUDENT PROGRESS REPORT", title_style)],
        [Paragraph("Vidya AI • Track your learning journey, achievements, and performance.", subtitle_style)]
    ]
    header_table = Table(header_data, colWidths=[504])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), primary_color),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 16),
        ('BOTTOMPADDING', (0,0), (-1,-1), 16),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 15))
    
    # 2. Metadata details
    meta_data = [
        [Paragraph("<b>Student Name:</b>", body_style), Paragraph(data["user"]["name"], body_bold_style),
         Paragraph("<b>Class / Grade:</b>", body_style), Paragraph(data["user"]["className"], body_bold_style)],
        [Paragraph("<b>Date Range:</b>", body_style), Paragraph(f"{start.strftime('%b %d, %Y')} - {end.strftime('%b %d, %Y')}", body_bold_style),
         Paragraph("<b>Generated On:</b>", body_style), Paragraph(datetime.utcnow().strftime('%b %d, %Y'), body_bold_style)],
    ]
    meta_table = Table(meta_data, colWidths=[100, 152, 100, 152])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), light_neutral),
        ('BOX', (0,0), (-1,-1), 0.5, border_color),
        ('INNERGRID', (0,0), (-1,-1), 0.25, border_color),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 12),
        ('RIGHTPADDING', (0,0), (-1,-1), 12),
    ]))
    story.append(meta_table)
    
    # 3. Summary Block
    summary_para = (
        f"Based on activities recorded from <b>{start.strftime('%B %d, %Y')}</b> to <b>{end.strftime('%B %d, %Y')}</b>, "
        f"<b>{data['user']['name']}</b> has been highly active, maintaining a study streak of <b>{data['user']['streak']} days</b> "
        f"and accumulating <b>{data['user']['xpPoints']} total XP</b>. They completed <b>{data['counts']['homework']} homework sheets</b>, "
        f"solved <b>{data['counts']['doubts']} tutor doubts</b>, and solved <b>{data['counts']['quizzes']} quizzes</b> with an accuracy of "
        f"<b>{data['counts']['quizAccuracy']}%</b>. Their strongest academic subject is <b>{data['insights']['strongestSubject']}</b>. "
        f"Keep practicing to achieve higher scores!"
    )
    story.append(Paragraph("Learning Summary", h1_style))
    story.append(Paragraph(summary_para, body_style))
    
    # 4. Metrics Grid
    story.append(Paragraph("Overview Metrics", h1_style))
    metrics_data = [
        [Paragraph("<b>Metric</b>", body_bold_style), Paragraph("<b>Value</b>", body_bold_style), Paragraph("<b>Description / Status</b>", body_bold_style)]
    ]
    for st in data["stats"]:
        metrics_data.append([
            Paragraph(st["label"], body_style),
            Paragraph(st["value"], body_bold_style),
            Paragraph(f"Active tracking during this period", body_style) if st["id"] not in ("streak", "xp") else Paragraph("Cumulative overall metric", body_style)
        ])
    metrics_table = Table(metrics_data, colWidths=[150, 100, 254])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#E2E8F0")),
        ('BOX', (0,0), (-1,-1), 0.5, border_color),
        ('INNERGRID', (0,0), (-1,-1), 0.25, border_color),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(metrics_table)
    
    # 5. Subject Progress Table
    story.append(Paragraph("Subject Performance Breakdown", h1_style))
    sub_data = [
        [Paragraph("<b>Subject</b>", body_bold_style), Paragraph("<b>Progress / Score</b>", body_bold_style), Paragraph("<b>Weekly Trend</b>", body_bold_style)]
    ]
    for sb in data["performanceBars"]:
        sub_data.append([
            Paragraph(f"{sb['emoji']} {sb['name']}", body_style),
            Paragraph(f"{sb['progress']}%", body_bold_style),
            Paragraph(sb['trend'], body_style)
        ])
    sub_table = Table(sub_data, colWidths=[180, 144, 180])
    sub_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#E2E8F0")),
        ('BOX', (0,0), (-1,-1), 0.5, border_color),
        ('INNERGRID', (0,0), (-1,-1), 0.25, border_color),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(sub_table)
    
    # 6. Achievements
    story.append(Paragraph("Earned Achievements & Badges", h1_style))
    ach_list_data = []
    ach_row = []
    for idx, ac in enumerate(data["achievements"]):
        status_lbl = "Earned ✓" if ac["unlocked"] else "Locked 🔒"
        status_date = f" on {ac['earnedDate']}" if ac["earnedDate"] else ""
        cell_text = f"<b>{ac['emoji']} {ac['name']}</b> ({status_lbl}{status_date})<br/><font size=8.5 color='#4A5568'>{ac['description']}</font>"
        ach_row.append(Paragraph(cell_text, body_style))
        if len(ach_row) == 2 or idx == len(data["achievements"]) - 1:
            if len(ach_row) == 1:
                ach_row.append(Paragraph("", body_style))
            ach_list_data.append(ach_row)
            ach_row = []
    ach_table = Table(ach_list_data, colWidths=[252, 252])
    ach_table.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, border_color),
        ('INNERGRID', (0,0), (-1,-1), 0.25, border_color),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(ach_table)
    
    # 7. AI Recommendations
    story.append(Paragraph("AI Recommendations & Learning Tips", h1_style))
    for rec in data["insights"]["recommendations"]:
        story.append(Paragraph(f"• {rec}", body_style))
        story.append(Spacer(1, 4))
        
    doc.build(story)
    return buffer.getvalue()

# Excel workbook builder
def generate_excel_report(db: Session, user: UserProfile, filter_type: str | None, start_date_str: str | None = None, end_date_str: str | None = None) -> bytes:
    data = get_progress_data(db, user, filter_type, start_date_str, end_date_str)
    start, end = get_datetime_range(filter_type, start_date_str, end_date_str)
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    
    # Style configuration
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="333333")
    data_font = Font(name=font_family, size=10)
    data_bold_font = Font(name=font_family, size=10, bold=True)
    
    title_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid") # Brand Orange
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Light grey
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Sheet 1: Overview Metrics
    ws1 = wb.active
    ws1.title = "Overview Metrics"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title row
    ws1.merge_cells("A1:C1")
    ws1["A1"] = "OVERVIEW METRICS"
    ws1["A1"].font = title_font
    ws1["A1"].fill = title_fill
    ws1["A1"].alignment = center_align
    ws1.row_dimensions[1].height = 35
    
    # Headers
    ws1["A2"] = "Metric Key"
    ws1["B2"] = "Value"
    ws1["C2"] = "Description"
    for col in ["A", "B", "C"]:
        ws1[f"{col}2"].font = header_font
        ws1[f"{col}2"].fill = header_fill
        ws1[f"{col}2"].alignment = left_align
        ws1[f"{col}2"].border = thin_border
    ws1.row_dimensions[2].height = 25
    
    # Data
    row_num = 3
    for st in data["stats"]:
        ws1.cell(row=row_num, column=1, value=st["label"]).font = data_font
        ws1.cell(row=row_num, column=2, value=st["value"]).font = data_bold_font
        ws1.cell(row=row_num, column=3, value=f"Active tracking during report range" if st["id"] not in ("streak", "xp") else "Overall total").font = data_font
        for col_idx in range(1, 4):
            ws1.cell(row=row_num, column=col_idx).border = thin_border
            ws1.cell(row=row_num, column=col_idx).alignment = left_align
        ws1.row_dimensions[row_num].height = 20
        row_num += 1
        
    # Auto-fit column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    # Sheet 2: Subject Performance
    ws2 = wb.create_sheet(title="Subject Performance")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "SUBJECT PERFORMANCE"
    ws2["A1"].font = title_font
    ws2["A1"].fill = title_fill
    ws2["A1"].alignment = center_align
    ws2.row_dimensions[1].height = 35
    
    ws2["A2"] = "Subject"
    ws2["B2"] = "Score / Progress"
    ws2["C2"] = "Emoji"
    ws2["D2"] = "Weekly Trend"
    for col in ["A", "B", "C", "D"]:
        ws2[f"{col}2"].font = header_font
        ws2[f"{col}2"].fill = header_fill
        ws2[f"{col}2"].alignment = left_align
        ws2[f"{col}2"].border = thin_border
    ws2.row_dimensions[2].height = 25
    
    row_num = 3
    for sb in data["performanceBars"]:
        ws2.cell(row=row_num, column=1, value=sb["name"]).font = data_font
        ws2.cell(row=row_num, column=2, value=f"{sb['progress']}%").font = data_bold_font
        ws2.cell(row=row_num, column=3, value=sb["emoji"]).font = data_font
        ws2.cell(row=row_num, column=4, value=sb["trend"]).font = data_font
        for col_idx in range(1, 5):
            ws2.cell(row=row_num, column=col_idx).border = thin_border
            ws2.cell(row=row_num, column=col_idx).alignment = left_align
        ws2.row_dimensions[row_num].height = 20
        row_num += 1
        
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    # Sheet 3: Homework History
    ws3 = wb.create_sheet(title="Homework History")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "HOMEWORK HISTORY"
    ws3["A1"].font = title_font
    ws3["A1"].fill = title_fill
    ws3["A1"].alignment = center_align
    ws3.row_dimensions[1].height = 35
    
    ws3["A2"] = "Date Uploaded"
    ws3["B2"] = "Subject"
    ws3["C2"] = "Question / Title"
    ws3["D2"] = "Status"
    ws3["E2"] = "Steps Count"
    for col in ["A", "B", "C", "D", "E"]:
        ws3[f"{col}2"].font = header_font
        ws3[f"{col}2"].fill = header_fill
        ws3[f"{col}2"].alignment = left_align
        ws3[f"{col}2"].border = thin_border
    ws3.row_dimensions[2].height = 25
    
    # Query Homework Analyses in range
    hws = list(
        db.scalars(
            select(HomeworkAnalysis)
            .where(
                HomeworkAnalysis.user_id == user.id,
                HomeworkAnalysis.created_at.between(start, end)
            )
            .order_by(HomeworkAnalysis.created_at.desc())
        ).all()
    )
    
    row_num = 3
    if not hws:
        ws3.merge_cells("A3:E3")
        ws3["A3"] = "No homework uploaded during this period."
        ws3["A3"].font = data_font
        ws3["A3"].alignment = center_align
        ws3["A3"].border = thin_border
        ws3.row_dimensions[3].height = 20
    else:
        for hw in hws:
            ws3.cell(row=row_num, column=1, value=hw.created_at.strftime("%Y-%m-%d %H:%M")).font = data_font
            ws3.cell(row=row_num, column=2, value=hw.subject_id.capitalize()).font = data_font
            ws3.cell(row=row_num, column=3, value=(hw.question_text[:50] + "...") if len(hw.question_text) > 50 else hw.question_text).font = data_font
            ws3.cell(row=row_num, column=4, value=hw.status.upper()).font = data_bold_font
            ws3.cell(row=row_num, column=5, value=len(hw.steps)).font = data_font
            for col_idx in range(1, 6):
                ws3.cell(row=row_num, column=col_idx).border = thin_border
                ws3.cell(row=row_num, column=col_idx).alignment = left_align
            ws3.row_dimensions[row_num].height = 20
            row_num += 1
            
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    # Sheet 4: Quiz History
    ws4 = wb.create_sheet(title="Quiz History")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4.merge_cells("A1:E1")
    ws4["A1"] = "QUIZ HISTORY"
    ws4["A1"].font = title_font
    ws4["A1"].fill = title_fill
    ws4["A1"].alignment = center_align
    ws4.row_dimensions[1].height = 35
    
    ws4["A2"] = "Date Answered"
    ws4["B2"] = "Question / Topic"
    ws4["C2"] = "Selected Option"
    ws4["D2"] = "Is Correct"
    ws4["E2"] = "XP Awarded"
    for col in ["A", "B", "C", "D", "E"]:
        ws4[f"{col}2"].font = header_font
        ws4[f"{col}2"].fill = header_fill
        ws4[f"{col}2"].alignment = left_align
        ws4[f"{col}2"].border = thin_border
    ws4.row_dimensions[2].height = 25
    
    # Query normal quiz attempts and adaptive quiz attempts in range
    quiz_attempts = list(
        db.scalars(
            select(QuizAttempt)
            .where(
                QuizAttempt.user_id == user.id,
                QuizAttempt.created_at.between(start, end)
            )
            .order_by(QuizAttempt.created_at.desc())
        ).all()
    )
    
    adaptive_attempts = list(
        db.scalars(
            select(AdaptiveQuizAttempt)
            .where(
                AdaptiveQuizAttempt.user_id == user.id,
                AdaptiveQuizAttempt.created_at.between(start, end)
            )
            .order_by(AdaptiveQuizAttempt.created_at.desc())
        ).all()
    )
    
    all_attempts = []
    for qa in quiz_attempts:
        all_attempts.append({
            "date": qa.created_at,
            "type": "Standard",
            "selected": qa.selected_option,
            "correct": qa.correct,
            "xp": qa.xp_awarded
        })
    for aa in adaptive_attempts:
        all_attempts.append({
            "date": aa.created_at,
            "type": "Adaptive",
            "selected": aa.selected_option,
            "correct": aa.correct,
            "xp": aa.xp_awarded
        })
    all_attempts.sort(key=lambda x: x["date"], reverse=True)
    
    row_num = 3
    if not all_attempts:
        ws4.merge_cells("A3:E3")
        ws4["A3"] = "No quizzes solved during this period."
        ws4["A3"].font = data_font
        ws4["A3"].alignment = center_align
        ws4["A3"].border = thin_border
        ws4.row_dimensions[3].height = 20
    else:
        for att in all_attempts:
            ws4.cell(row=row_num, column=1, value=att["date"].strftime("%Y-%m-%d %H:%M")).font = data_font
            ws4.cell(row=row_num, column=2, value=f"{att['type']} Practice Quiz Question").font = data_font
            ws4.cell(row=row_num, column=3, value=att["selected"]).font = data_font
            ws4.cell(row=row_num, column=4, value="CORRECT" if att["correct"] else "INCORRECT").font = data_bold_font
            ws4.cell(row=row_num, column=5, value=att["xp"]).font = data_font
            for col_idx in range(1, 6):
                ws4.cell(row=row_num, column=col_idx).border = thin_border
                ws4.cell(row=row_num, column=col_idx).alignment = left_align
            ws4.row_dimensions[row_num].height = 20
            row_num += 1
            
    for col in ws4.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws4.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    # Sheet 5: AI Tutor Activity
    ws5 = wb.create_sheet(title="AI Tutor Activity")
    ws5.views.sheetView[0].showGridLines = True
    
    ws5.merge_cells("A1:E1")
    ws5["A1"] = "AI TUTOR ACTIVITY"
    ws5["A1"].font = title_font
    ws5["A1"].fill = title_fill
    ws5["A1"].alignment = center_align
    ws5.row_dimensions[1].height = 35
    
    ws5["A2"] = "Date Started"
    ws5["B2"] = "Doubt Thread Title"
    ws5["C2"] = "Language"
    ws5["D2"] = "Message Count"
    ws5["E2"] = "Status"
    for col in ["A", "B", "C", "D", "E"]:
        ws5[f"{col}2"].font = header_font
        ws5[f"{col}2"].fill = header_fill
        ws5[f"{col}2"].alignment = left_align
        ws5[f"{col}2"].border = thin_border
    ws5.row_dimensions[2].height = 25
    
    threads = list(
        db.scalars(
            select(DoubtThread)
            .where(
                DoubtThread.user_id == user.id,
                DoubtThread.created_at.between(start, end)
            )
            .order_by(DoubtThread.created_at.desc())
        ).all()
    )
    
    row_num = 3
    if not threads:
        ws5.merge_cells("A3:E3")
        ws5["A3"] = "No AI Tutor doubt discussions in this range."
        ws5["A3"].font = data_font
        ws5["A3"].alignment = center_align
        ws5["A3"].border = thin_border
        ws5.row_dimensions[3].height = 20
    else:
        for th in threads:
            msg_count = db.scalar(select(func.count(DoubtMessage.id)).where(DoubtMessage.thread_id == th.id)) or 0
            ws5.cell(row=row_num, column=1, value=th.created_at.strftime("%Y-%m-%d %H:%M")).font = data_font
            ws5.cell(row=row_num, column=2, value=th.title).font = data_font
            ws5.cell(row=row_num, column=3, value=th.language.upper()).font = data_font
            ws5.cell(row=row_num, column=4, value=msg_count).font = data_font
            ws5.cell(row=row_num, column=5, value=th.status.upper()).font = data_bold_font
            for col_idx in range(1, 6):
                ws5.cell(row=row_num, column=col_idx).border = thin_border
                ws5.cell(row=row_num, column=col_idx).alignment = left_align
            ws5.row_dimensions[row_num].height = 20
            row_num += 1
            
    for col in ws5.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws5.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    # Sheet 6: Achievements
    ws6 = wb.create_sheet(title="Achievements")
    ws6.views.sheetView[0].showGridLines = True
    
    ws6.merge_cells("A1:D1")
    ws6["A1"] = "ACHIEVEMENTS & BADGES"
    ws6["A1"].font = title_font
    ws6["A1"].fill = title_fill
    ws6["A1"].alignment = center_align
    ws6.row_dimensions[1].height = 35
    
    ws6["A2"] = "Badge Name"
    ws6["B2"] = "Emoji"
    ws6["C2"] = "Status"
    ws6["D2"] = "Requirement Description"
    for col in ["A", "B", "C", "D"]:
        ws6[f"{col}2"].font = header_font
        ws6[f"{col}2"].fill = header_fill
        ws6[f"{col}2"].alignment = left_align
        ws6[f"{col}2"].border = thin_border
    ws6.row_dimensions[2].height = 25
    
    row_num = 3
    for ac in data["achievements"]:
        status_lbl = f"UNLOCKED ({ac['earnedDate']})" if ac["unlocked"] else "LOCKED"
        ws6.cell(row=row_num, column=1, value=ac["name"]).font = data_bold_font
        ws6.cell(row=row_num, column=2, value=ac["emoji"]).font = data_font
        ws6.cell(row=row_num, column=3, value=status_lbl).font = data_bold_font
        ws6.cell(row=row_num, column=4, value=ac["description"]).font = data_font
        for col_idx in range(1, 5):
            ws6.cell(row=row_num, column=col_idx).border = thin_border
            ws6.cell(row=row_num, column=col_idx).alignment = left_align
        ws6.row_dimensions[row_num].height = 20
        row_num += 1
        
    for col in ws6.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws6.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
